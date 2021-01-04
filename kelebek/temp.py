from openpyxl import load_workbook

class sinif():
    # def __init__(self,sinif_adi,ogr_sayisi):
    def __init__(self):
        self.sinif_ogrencileri = []
    def sinif_tanimla(self,sinif_adi):
        self.sinif_adi = sinif_adi
    def ogr_sayi(self,ogr_sayisi):
        self.ogr_sayisi = ogr_sayisi

    def ogrenci_ekle(self,ogrenci):
        self.sinif_ogrencileri.append(ogrenci)

    def ogrenci_listesi(self):
        for ogrenci in self.sinif_ogrencileri:
            print("öğrenci no:{}, öğrenci adı : {}, öğrenci soyadı : {}".format(ogrenci[0],ogrenci[1],ogrenci[2]))


wb = load_workbook(filename='okullistesi.xlsx', read_only=True)
ws = wb['Sheet1']
# print(ws.max_row)
ogrenciler = []
sinif_sayilari = []
for deger in range (2,ws.max_row):
    hdegeri = "P"+str(deger-1)
    if ws[hdegeri].value != None:
        smevcut = ws[hdegeri].value
        sinif_sayilari.append(smevcut)
        # print(ogrenciler)


# ssayisi = ws[satir].value
# print(ssayisi)
#  hucreler A12 den başlıyor

baslangic = 13
siniflar = ["sinif_9a","sinif_9b","sinif_9c","sinif_10a","sinif_10b",
            "sinif_10c","sinif_11a","sinif_11b","sinif_11c","sinif_11d",
            "sinif_12a","sinif_12b","sinif_12c","sinif_12d","sinif_12e"]
ssiniflar = iter(siniflar)
fsinif = {}
for ogrenci in siniflar:
    exec(ogrenci + " = sinif()")

for ilerleme\


        in sinif_sayilari:
    x_sinif = str(next(ssiniflar))
    # o_sinif = x_sinif.__repr__()
    fsinif[x_sinif]=ilerleme
    o_sinif = eval(x_sinif)
    # o_sinif = sinif()
    o_sinif.sinif_tanimla(x_sinif)
    o_sinif.ogr_sayi(ilerleme)
    # tanimlama = "{}={}".format(o_sinif,sinif())
    # exec(x_sinif +" = o_sinif()")


    # sinif_tanimlama = "{} = {}".format(x_sinif,o_sinif)
    # exec(sinif_tanimlama)
    # eval(repr(x_sinif),{x_sinif: o_sinif})
    # print(a)
    # print(o_sinif.sinif_adi)
    # ogrenciler.append([next(ssiniflar),ilerleme])
    for deger in range(baslangic,baslangic+ilerleme*2,2):
        ssatir = ws[deger]
    # print(ssatir[1])
    # okunacak satırlar 2,5,10

        ogrenci = []
        for cell in ssatir[1],ssatir[4],ssatir[9]:
        # print(cell.value)
            ogrenci.append(cell.value)
        o_sinif.ogrenci_ekle(ogrenci)
    # o_sinif.ogrenci_listesi()
    baslangic = baslangic + ilerleme*2+ 1

# asd = "{}{}".format(str(siniflar[0]),".ogrenci_listesi()")
# print(asd)
# # eval("sinif(asd).ogrenci_listesi()",{"k":asd})
# exec(asd)
sinif_9a.ogrenci_listesi()
k = sinif_12a.sinif_adi
print(fsinif[k])
