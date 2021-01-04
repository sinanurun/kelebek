from openpyxl import Workbook
wb = Workbook(write_only=True)
ws = wb.create_sheet()
ws2 = wb.create_sheet()


for irow in range(100):
    ws.append(['%d' % i for i in range(200)])

wb.save('new_big_file.xlsx')