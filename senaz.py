from openpyxl import load_workbook
from random import *

class sinif():
    def __init__(self):
        self.sinif_ogrencileri = []
    def sinif_tanimla(self,sinif_adi):
        self.sinif_adi = sinif_adi
    def ogr_sayi(self,ogr_sayisi):
        self.ogr_sayisi = ogr_sayisi
    def ogrenci_sayisi(self):
        return len(self.sinif_ogrencileri)
    def ogrenci_ekle(self,ogrenci):
        self.sinif_ogrencileri.append(ogrenci)
    def gidilecek_sinif_ekle(self,ogr_no):
        self.gidilecek_sinif=ogr_no
    def gsl(self):
        for ogrenci in self.sinif_ogrencileri:
            print("öğrenci no:{}, öğrenci adı : {}, öğrenci soyadı : {}, öğrencinin sinifi : {}, Öğrencinin Gideceği sinif : {}".format(ogrenci[0],ogrenci[1],ogrenci[2],ogrenci[3],ogrenci[4]))
    def ogrenci_listesi(self):
        for ogrenci in self.sinif_ogrencileri:
            print("öğrenci no:{:10}, öğrenci adı : {:20}, öğrenci soyadı : {:20}, öğrencinin sinifi : {:10}".format(ogrenci[0],ogrenci[1],ogrenci[2],ogrenci[3]))
        return self.sinif_ogrencileri

class f_sinif():
    def __init__(self):
        self.sinif_ogrencileri = []
    def sinif_tanimla(self,sinif_adi):
        self.sinif_adi = sinif_adi
    def ogr_sayi(self,ogr_sayisi):
        self.ogr_sayisi = len(self.sinif_ogrencileri)
    def ogrenci_ekle(self,ogrenci):
        self.sinif_ogrencileri.append(ogrenci)
    def gidilecek_sinif_ekle(self,ogr_no):
        self.gidilecek_sinif=ogr_no
    def gsl(self):
        for ogrenci in self.sinif_ogrencileri:
            print("öğrenci no:{}, öğrenci adı : {}, öğrenci soyadı : {}, öğrencinin sinifi : {}, Öğrencinin Gideceği sinif : {}".format(ogrenci[0],ogrenci[1],ogrenci[2],ogrenci[3],ogrenci[4]))
    def ogrenci_listesi(self):
        for ogrenci in self.sinif_ogrencileri:
            print("öğrenci no:{}, öğrenci adı : {:3}, öğrenci soyadı : {}, öğrencinin sinifi : {}".format(ogrenci[0],ogrenci[1],ogrenci[2],ogrenci[3]))
        return self.sinif_ogrencileri

wb = load_workbook(filename='okullistesi.xlsx', read_only=True)
ws = wb['Sheet1']
# print(ws.max_row)
ogrenciler = []
sinif_sayilari = []

for deger in range(2,ws.max_row):
    hdegeri = "P"+str(deger-1)
    if ws[hdegeri].value != None:
        smevcut = ws[hdegeri].value
        sinif_sayilari.append(smevcut)

baslangic = 13
siniflar = ["sinif_9a","sinif_9b","sinif_9c","sinif_10a","sinif_10b",
            "sinif_10c","sinif_11a","sinif_11b","sinif_11c","sinif_11d",
            "sinif_12a","sinif_12b","sinif_12c","sinif_12d","sinif_12e"]
ssiniflar = iter(siniflar)
fsinif = {}
for ogrenci in siniflar:
    exec(ogrenci + " = sinif()")
    exec("f_" + ogrenci + " = f_sinif()")
# print(sum(sinif_sayilari))
for ilerleme in sinif_sayilari:
    x_sinif = str(next(ssiniflar))
    fsinif[x_sinif]=ilerleme
    o_sinif = eval(x_sinif)
    o_sinif.sinif_tanimla(x_sinif)
    o_sinif.ogr_sayi(ilerleme)

    for deger in range(baslangic,baslangic+ilerleme*2,2):
        ssatir = ws[deger]
        ogrenci = []
        for cell in ssatir[1],ssatir[4],ssatir[9]:
            ogrenci.append(cell.value)
        ogrenci.append(x_sinif)
        o_sinif.ogrenci_ekle(ogrenci)

    baslangic = baslangic + ilerleme*2+ 1
fsiniflar = iter(siniflar)
for ilerleme in sinif_sayilari:
    x_sinif = "f_"+str(next(fsiniflar))
    # fsinif[x_sinif]=ilerleme
    o_sinif = eval(x_sinif)
    o_sinif.sinif_tanimla(x_sinif)
    o_sinif.ogr_sayi(ilerleme)

# esinif eski ysinif yeni sinif
def karisik_listele(esinif,ysiniff):
    esinif = eval(esinif)
    ysinif = eval("f_"+ysiniff)
    sayi = esinif.ogrenci_sayisi()
    rast = randrange(sayi)
    if len(ysinif.sinif_ogrencileri) < fsinif[ysiniff]:
        ysinif.ogrenci_ekle(esinif.sinif_ogrencileri[rast])
        print(esinif.sinif_ogrencileri[rast]," = >",ysiniff)
        esinif.sinif_ogrencileri.pop(rast)
        global toplam
        toplam += 1
        return True
    else:
        return False
z = fsinif
toplam = 0

while toplam < sum(sinif_sayilari):
    kontrol = 0
    for a in siniflar[:3], siniflar[6:10], siniflar[3:6], siniflar[10:]:
        for zr in a:
            print(a, fsinif[zr])
            o_sinif = eval(zr)

            mevcut = 1

            while len(o_sinif.sinif_ogrencileri) > 0:
                print(mevcut, end=" ")
                k = siniflar[(kontrol % 15)]
                b = k
                if karisik_listele(zr, b):
                    mevcut = mevcut + 1
                kontrol += 1

print("devam")
ert = 0
for a in z:
    print(a,"=>",z[a])
    y_sinif = eval("f_"+a)
    print(y_sinif.sinif_adi, len(y_sinif.sinif_ogrencileri))
    # ert +=len(y_sinif.ogrenci_listesi())
    y_sinif.ogrenci_listesi()
print(ert)
